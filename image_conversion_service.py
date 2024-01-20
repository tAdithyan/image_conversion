import pytesseract
from PIL import Image
import openpyxl
import os
from flask import Flask, request, jsonify, send_file

app = Flask(__name__)
# Directory to store uploaded files
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
# Ensure the upload folder exists
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
      # Step 1: OCR with Tesseract
    def image_to_excel(input_image, output_excel):
        try:
            # Perform OCR on the image
            ocr_result = pytesseract.image_to_string(Image.open(input_image))
            # Step 2: Text Parsing and Table Extraction (Simplified)

            lines = ocr_result.split('\n')
            column_headers = lines[0].split('\t')
            data = [line.split('\t') for line in lines[1:]]
            # Step 3: Generate Excel File
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            # Write column headers
            for col_num, header in enumerate(column_headers, start=1):
                worksheet.cell(row=1, column=col_num, value=header)
            # Write data
            for row_num, row_data in enumerate(data, start=2):
                for col_num, cell_value in enumerate(row_data, start=1):
                    worksheet.cell(row=row_num, column=col_num, value=cell_value)
            # Save the Excel file
            workbook.save(output_excel)
            return "Image to Excel conversion completed successfully."
        except Exception as e:
            return "Error during Image to Excel conversion: {str(e)}"


    # Step 4: Image to Different Image Types
def image_to_different_types(input_image, output_image, image_type):
        try:
            # Open and save the image in the specified format (JPEG, PNG, or GIF)
            img = Image.open(input_image)
            img.save(output_image, image_type)
        except Exception as e:
            return f"Error during Image to {image_type} conversion: {str(e)}"


    # Step 5: Image to PDF
def image_to_pdf(input_image, output_pdf):
        try:
            # Open and save the image as PDF
            img = Image.open(input_image)
            Image.Image.show(img)


            img.save(output_pdf, "PDF", resolution=100.0)
        except Exception as e:
            return f"Error during Image to PDF conversion: {str(e)}"


# API Endpoints


# Image to Excel Conversion API Endpoint
@app.route('/image_to_excel', methods=['POST'])
def image_to_excel_endpoint():
    try:
        # Receive the uploaded image file
        uploaded_file = request.files['file']
        input_image = os.path.join(app.config['UPLOAD_FOLDER'], 'input_image.png')
        uploaded_file.save(input_image)
        # Perform Image to Excel Conversion
        output_excel = os.path.join(app.config['UPLOAD_FOLDER'], 'output.xlsx')
        result = image_to_excel(input_image, output_excel)
        return result
    except Exception as e:
        return jsonify({'error': str(e)})


# Image to Different Image Types API Endpoint
@app.route('/image_to_different_types', methods=['POST'])
def image_to_different_types_endpoint():
    try:
        # Receive the uploaded image file
        uploaded_file = request.files['file']
        input_image = os.path.join(app.config['UPLOAD_FOLDER'], 'input_image.png')
        uploaded_file.save(input_image)
        # Perform Image to Different Types Conversion
        output_images = []
        for image_type in ['JPEG', 'PNG', 'GIF']:
            output_image = os.path.join(app.config['UPLOAD_FOLDER'], f'output.{image_type.lower()}')
            image_to_different_types(input_image, output_image, image_type)
            output_images.append(output_image)
        return jsonify({'output_images': output_images})
    except Exception as e:
        print("error dueeing api end point")
        return jsonify({'error': str(e)})


# Image to PDF Conversion API Endpoint
@app.route('/image_to_pdf', methods=['POST'])
def image_to_pdf_endpoint():
    try:
        # Receive the uploaded image file
        uploaded_file = request.files['file']
        input_image = os.path.join(app.config['UPLOAD_FOLDER'], 'uploads/input_image.png')
        uploaded_file.save(input_image)
        # Perform Image to PDF Conversion
        output_pdf = os.path.join(app.config['UPLOAD_FOLDER'], 'output.pdf')
        image_to_pdf(input_image, output_pdf)
        return send_file(output_pdf)
    except Exception as e:
        print(f"Error during Image to PDF conversion: {str(e)}")
        return jsonify({'error': str(e)})


if __name__ == "__main__":
    app.run(debug=True, threaded=True)
